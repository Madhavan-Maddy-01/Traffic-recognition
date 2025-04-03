import cv2
import numpy as np
import pytesseract
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

class SpeedViolationSystem:
    def __init__(self):
        # Initialize pytesseract path (Windows)
        pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        self.excel_file = 'speed_violations.xlsx'
        self.initialize_excel()

    def initialize_excel(self):
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            headers = ['Date', 'Time', 'License Plate', 'Speed (km/h)', 'Location']
            ws.append(headers)
            wb.save(self.excel_file)

    def detect_license_plate(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blur, 50, 150)
        
        contours, _ = cv2.findContours(edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10]
        
        plate_contour = None
        for contour in contours:
            perimeter = cv2.arcLength(contour, True)
            approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
            if len(approx) == 4:
                plate_contour = approx
                break
                
        if plate_contour is not None:
            x, y, w, h = cv2.boundingRect(plate_contour)
            plate_img = gray[y:y+h, x:x+w]
            return plate_img, (x, y, w, h)
        return None, None

    def read_license_plate(self, plate_img):
        if plate_img is not None:
            text = pytesseract.image_to_string(plate_img, config='--psm 8')
            return text.strip()
        return None

    def record_violation(self, license_plate, speed, location):
        now = datetime.now()
        wb = load_workbook(self.excel_file)
        ws = wb.active
        ws.append([
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            license_plate,
            speed,
            location
        ])
        wb.save(self.excel_file)

    def process_video(self, video_path, speed_limit, location):
        cap = cv2.VideoCapture(video_path)
        
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            plate_img, coords = self.detect_license_plate(frame)
            if plate_img is not None:
                license_plate = self.read_license_plate(plate_img)
                if license_plate:
                    # Simulated speed (in real implementation, this would come from speed detection hardware)
                    simulated_speed = np.random.randint(speed_limit - 20, speed_limit + 30)
                    
                    if simulated_speed > speed_limit:
                        self.record_violation(license_plate, simulated_speed, location)
                        
                        # Draw rectangle around plate
                        x, y, w, h = coords
                        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                        cv2.putText(frame, f"Plate: {license_plate}", (x, y-10),
                                  cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                        cv2.putText(frame, f"Speed: {simulated_speed} km/h", (x, y+h+25),
                                  cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

            cv2.imshow('Speed Violation Detection', frame)
            
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

# Example usage
if __name__ == "__main__":
    system = SpeedViolationSystem()
    system.process_video('video1.mp4', speed_limit=60, location='Main Street')